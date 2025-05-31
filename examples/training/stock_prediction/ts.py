from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from xgboost import XGBRegressor
from sklearn.metrics import mean_squared_error
import yfinance as yf
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Set up logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# List of energy stock symbols
ENERGY_STOCKS = [
    "XOM",
    "CVX",
    "COP",
    "EOG",
    "SLB",
    "PXD",
    "MPC",
    "VLO",
    "PSX",
    "OXY",
    "KMI",
    "WMB",
    "HAL",
    "DVN",
    "BKR",
]


def fetch_stock_data(symbol, start_date, end_date):
    """Fetch stock data from Yahoo Finance."""
    logging.info(f"Fetching stock data for {symbol}")
    stock = yf.Ticker(symbol)
    df = stock.history(start=start_date, end=end_date)
    return df


def prepare_features(df):
    """Prepare features for the model."""
    df["Returns"] = df["Close"].pct_change()
    df["MA_5"] = df["Close"].rolling(window=5).mean()
    df["MA_20"] = df["Close"].rolling(window=20).mean()
    df["RSI"] = calculate_rsi(df["Close"])
    df["Volume_Change"] = df["Volume"].pct_change()
    df.dropna(inplace=True)
    return df


def calculate_rsi(prices, period=14):
    """Calculate Relative Strength Index."""
    delta = prices.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def prepare_target(df, forecast_horizon=5):
    """Prepare target variable."""
    df["Target"] = (
        df["Close"]
        .pct_change(periods=forecast_horizon)
        .shift(-forecast_horizon)
    )
    df.dropna(inplace=True)
    return df


def train_model(X, y):
    """Train XGBoost model."""
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42
    )

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    model = XGBRegressor(
        n_estimators=1000, learning_rate=0.01, max_depth=5, random_state=42
    )
    model.fit(X_train_scaled, y_train)

    # Evaluate the model
    train_rmse = mean_squared_error(
        y_train, model.predict(X_train_scaled), squared=False
    )
    test_rmse = mean_squared_error(
        y_test, model.predict(X_test_scaled), squared=False
    )

    logging.info(f"Training RMSE: {train_rmse}")
    logging.info(f"Test RMSE: {test_rmse}")

    return model, scaler


def evaluate_stock(symbol, start_date, end_date):
    """Evaluate a single stock and return its predicted performance."""
    df = fetch_stock_data(symbol, start_date, end_date)
    df = prepare_features(df)
    df = prepare_target(df)

    features = ["Returns", "MA_5", "MA_20", "RSI", "Volume_Change"]
    X = df[features]
    y = df["Target"]

    model, scaler = train_model(X, y)

    # Predict future performance
    latest_data = X.iloc[-1].values.reshape(1, -1)
    latest_data_scaled = scaler.transform(latest_data)
    predicted_return = model.predict(latest_data_scaled)[0]

    return predicted_return


def save_results_to_excel(
    stock_predictions, filename="energy_stocks_analysis.xlsx"
):
    """Save the stock predictions to an Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Predictions"

    # Add headers
    headers = ["Rank", "Symbol", "Predicted 5-day Return"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Sort stocks by predicted return (descending order)
    sorted_predictions = sorted(
        stock_predictions, key=lambda x: x[1], reverse=True
    )

    # Add data
    for rank, (symbol, predicted_return) in enumerate(
        sorted_predictions, start=1
    ):
        ws.append([rank, symbol, f"{predicted_return:.2%}"])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add timestamp
    ws["A" + str(ws.max_row + 2)] = (
        f"Analysis performed on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )

    # Save the workbook
    wb.save(filename)
    logging.info(f"Results saved to {filename}")


def main():
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365 * 2)  # 2 years of data

    stock_predictions = []
    for symbol in ENERGY_STOCKS:
        try:
            predicted_return = evaluate_stock(symbol, start_date, end_date)
            stock_predictions.append((symbol, predicted_return))
            logging.info(
                f"Predicted 5-day return for {symbol}: {predicted_return:.2%}"
            )
        except Exception as e:
            logging.error(f"Error processing {symbol}: {str(e)}")

    # Sort stocks by predicted return (descending order)
    top_stocks = sorted(stock_predictions, key=lambda x: x[1], reverse=True)[
        :10
    ]

    logging.info("Top 10 energy stocks to buy right now:")
    for rank, (symbol, predicted_return) in enumerate(top_stocks, 1):
        logging.info(
            f"{rank}. {symbol}: Predicted 5-day return: {predicted_return:.2%}"
        )

    # Save results to Excel
    save_results_to_excel(stock_predictions)


if __name__ == "__main__":
    main()
